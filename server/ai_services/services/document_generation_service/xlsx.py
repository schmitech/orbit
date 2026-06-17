import io
from typing import Dict, Any

from .base import BaseRenderer


class XlsxRenderer(BaseRenderer):
    """Render a document spec to XLSX bytes using openpyxl."""

    @staticmethod
    def _xlsx_coerce(val) -> object:
        """Return a native int/float when val looks numeric, otherwise str."""
        s = str(val).strip()
        try:
            if '.' in s or 'e' in s.lower():
                return float(s)
            return int(s)
        except ValueError:
            return s

    def render(self, spec: Dict[str, Any]) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        header_color      = self._get('xlsx', 'colors', 'header_bg',   default='4472C4')
        alt_color         = self._get('xlsx', 'colors', 'alt_row_bg',  default='EBF0FA')
        header_text_color = self._get('xlsx', 'colors', 'header_text', default='FFFFFF')
        grid_color        = self._get('xlsx', 'colors', 'grid',        default='D9D9D9')
        font_body         = self._get('xlsx', 'fonts',  'body',        default='Calibri')
        font_header       = self._get('xlsx', 'fonts',  'header',      default='Calibri')
        body_size         = self._get('xlsx', 'fonts',  'body_size',   default=11)
        header_size       = self._get('xlsx', 'fonts',  'header_size', default=11)
        summary_col_width = self._get('xlsx', 'summary_col_width',     default=80)
        max_col_width     = self._get('xlsx', 'max_col_width',         default=50)
        min_col_width     = self._get('xlsx', 'min_col_width',         default=10)
        freeze_header     = self._get('xlsx', 'freeze_header_row',     default=True)
        use_auto_filter   = self._get('xlsx', 'auto_filter',           default=True)
        use_borders       = self._get('xlsx', 'borders',               default=True)

        wb = Workbook()
        wb.remove(wb.active)

        header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        alt_fill    = PatternFill(start_color=alt_color,    end_color=alt_color,    fill_type="solid")
        thin_side   = Side(style='thin', color=grid_color)
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        # Summary sheet
        ws = wb.create_sheet("Summary")
        ws["A1"] = spec.get("title", "Document")
        ws["A1"].font = Font(name=font_header, size=16, bold=True)

        meta_line = self._build_meta_line(spec)
        if meta_line:
            ws["A2"] = meta_line
            ws["A2"].font = Font(name=font_body, italic=True)
        ws.column_dimensions["A"].width = summary_col_width

        row = 4
        for section in spec.get("sections", []):
            if section.get("heading"):
                ws.cell(row=row, column=1, value=section["heading"]).font = Font(name=font_header, bold=True, size=12)
                row += 1
            if section.get("body"):
                ws.cell(row=row, column=1, value=section["body"]).font = Font(name=font_body, size=body_size)
                row += 1
            for point in section.get("bullet_points") or []:
                ws.cell(row=row, column=1, value=f"• {point}").font = Font(name=font_body, size=body_size)
                row += 1
            row += 1

        # Data sheets — one per section with a table
        for idx, section in enumerate(spec.get("sections", [])):
            table_data = section.get("table")
            if not table_data or len(table_data) < 1:
                continue
            sheet_name = (section.get("heading") or f"Sheet{idx + 1}")[:31]
            ws_data = wb.create_sheet(sheet_name)
            for r_idx, tbl_row in enumerate(table_data):
                for c_idx, val in enumerate(tbl_row):
                    coerced = self._xlsx_coerce(val) if r_idx > 0 else str(val)
                    cell = ws_data.cell(row=r_idx + 1, column=c_idx + 1, value=coerced)
                    if r_idx == 0:
                        cell.font      = Font(name=font_header, bold=True, size=header_size, color=header_text_color)
                        cell.fill      = header_fill
                        cell.alignment = Alignment(horizontal="center")
                    else:
                        cell.font = Font(name=font_body, size=body_size)
                        if r_idx % 2 == 0:
                            cell.fill = alt_fill
                    if use_borders:
                        cell.border = thin_border

            for col in ws_data.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                width = min(max_len + 4, max_col_width)
                width = max(width, min_col_width)
                ws_data.column_dimensions[col[0].column_letter].width = width

            if freeze_header:
                ws_data.freeze_panes = "A2"
            if use_auto_filter and table_data:
                ws_data.auto_filter.ref = ws_data.dimensions

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
