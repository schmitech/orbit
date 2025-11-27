"""
XLSX Processor

Handles Excel XLSX files using openpyxl.
"""

import logging
from typing import Dict, Any
from .base_processor import FileProcessor

logger = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook
    from io import BytesIO
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False
    logger.warning("openpyxl not available. XLSX processing disabled.")


class XLSXProcessor(FileProcessor):
    """
    Processor for Excel XLSX files.

    Supports: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet
    Requires: openpyxl
    """

    def supports_mime_type(self, mime_type: str) -> bool:
        """Check if this processor supports the MIME type."""
        xlsx_types = [
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'application/vnd.ms-excel',
        ]
        return XLSX_AVAILABLE and mime_type.lower() in xlsx_types

    async def extract_text(self, file_data: bytes, filename: str = None) -> str:
        """Extract text from XLSX."""
        if not XLSX_AVAILABLE:
            raise ImportError("openpyxl not available")

        logger.debug(f"XLSXProcessor.extract_text() called for file: {filename or 'unknown'} (using openpyxl)")

        text_parts = []

        try:
            wb = load_workbook(BytesIO(file_data), read_only=True, data_only=True)

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_texts = []

                # Get dimensions
                if sheet.max_row and sheet.max_column:
                    text_parts.append(f"--- Sheet: {sheet_name} ({sheet.max_row} rows x {sheet.max_column} columns) ---")

                    # Extract header row (first row)
                    header_row = []
                    for col in range(1, min(sheet.max_column + 1, 51)):  # Limit to 50 columns
                        cell = sheet.cell(row=1, column=col)
                        header_row.append(str(cell.value) if cell.value is not None else "")

                    if any(header_row):
                        text_parts.append("Headers: " + " | ".join(header_row))
                        text_parts.append("")

                    # Extract data rows (limit to first 100 rows for large files)
                    max_rows = min(sheet.max_row + 1, 101)
                    for row_num in range(2, max_rows):
                        row_texts = []
                        for col in range(1, min(sheet.max_column + 1, 51)):
                            cell = sheet.cell(row=row_num, column=col)
                            if cell.value is not None:
                                row_texts.append(str(cell.value))
                        if row_texts:
                            sheet_texts.append(" | ".join(row_texts))

                    if sheet_texts:
                        text_parts.extend(sheet_texts)

                    if sheet.max_row > 100:
                        text_parts.append(f"... (showing first 100 of {sheet.max_row} rows)")

                text_parts.append("")  # Empty line between sheets

            wb.close()
            return "\n".join(text_parts)

        except Exception as e:
            logger.error(f"Error processing XLSX: {e}")
            raise

    async def extract_metadata(self, file_data: bytes, filename: str = None) -> Dict[str, Any]:
        """Extract metadata from XLSX."""
        metadata = await super().extract_metadata(file_data, filename)

        if not XLSX_AVAILABLE:
            return metadata

        try:
            wb = load_workbook(BytesIO(file_data), read_only=True, data_only=True)

            # Count sheets and get their names
            sheet_names = wb.sheetnames
            sheet_count = len(sheet_names)

            # Get total row/column counts across all sheets
            total_rows = 0
            total_cols = 0
            sheet_info = []

            for sheet_name in sheet_names:
                sheet = wb[sheet_name]
                rows = sheet.max_row or 0
                cols = sheet.max_column or 0
                total_rows += rows
                total_cols = max(total_cols, cols)
                sheet_info.append({
                    'name': sheet_name,
                    'rows': rows,
                    'columns': cols
                })

            metadata.update({
                'sheet_count': sheet_count,
                'sheet_names': sheet_names,
                'sheets': sheet_info,
                'total_rows': total_rows,
                'max_columns': total_cols,
                'mime_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            })

            # Get workbook properties if available
            if wb.properties:
                props = wb.properties
                if props.title:
                    metadata['title'] = props.title
                if props.creator:
                    metadata['author'] = props.creator
                if props.created:
                    metadata['created'] = props.created.isoformat() if props.created else None
                if props.modified:
                    metadata['modified'] = props.modified.isoformat() if props.modified else None

            wb.close()

        except Exception as e:
            logger.warning(f"Error extracting XLSX metadata: {e}")

        return metadata
