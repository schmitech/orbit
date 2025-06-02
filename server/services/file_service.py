"""
File processing service for handling file uploads and text extraction
"""

import os
import io
import uuid
import logging
import hashlib
import mimetypes
from typing import Dict, Any, List, Optional, Union, BinaryIO
from pathlib import Path

# File processing libraries
try:
    import magic
    HAS_MAGIC = True
except ImportError:
    HAS_MAGIC = False
    
try:
    from pypdf import PdfReader
    HAS_PDF = True
except ImportError:
    HAS_PDF = False
    
try:
    from docx2python import docx2python
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False
    
try:
    import openpyxl
    import csv
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

logger = logging.getLogger(__name__)

class FileService:
    """Service for processing file uploads and extracting text content"""
    
    def __init__(self, config: Dict[str, Any], retriever_service=None):
        """
        Initialize the file service.
        
        Args:
            config: Application configuration
            retriever_service: Optional retriever service for storing documents
        """
        self.config = config
        self.retriever_service = retriever_service
        
        # File configuration
        file_config = config.get('file_upload', {})
        self.max_file_size = file_config.get('max_size_mb', 10) * 1024 * 1024  # Convert to bytes
        self.allowed_extensions = set(file_config.get('allowed_extensions', [
            '.txt', '.pdf', '.docx', '.doc', '.xlsx', '.xls', '.csv', '.md', '.json'
        ]))
        self.upload_dir = Path(file_config.get('upload_directory', 'uploads'))
        self.auto_store = file_config.get('auto_store_in_vector_db', True)
        self.chunk_size = file_config.get('chunk_size', 1000)
        self.chunk_overlap = file_config.get('chunk_overlap', 200)
        
        # Create upload directory
        self.upload_dir.mkdir(exist_ok=True)
        
        logger.info(f"FileService initialized with max_size: {self.max_file_size} bytes")
        
    async def process_file_upload(self, 
                                file_content: bytes, 
                                filename: str, 
                                collection_name: str = "default",
                                metadata: Dict[str, Any] = None) -> Dict[str, Any]:
        """
        Process an uploaded file and extract text content.
        
        Args:
            file_content: The file content as bytes
            filename: Original filename
            collection_name: Collection to store the document in
            metadata: Additional metadata for the document
            
        Returns:
            Dictionary with processing results
        """
        try:
            # Validate file
            validation_result = self._validate_file(file_content, filename)
            if not validation_result['valid']:
                return {
                    'success': False,
                    'error': validation_result['error'],
                    'file_id': None
                }
            
            # Generate unique file ID
            file_id = str(uuid.uuid4())
            file_hash = hashlib.sha256(file_content).hexdigest()
            
            # Detect MIME type
            mime_type = self._detect_mime_type(file_content, filename)
            
            # Extract text content
            extraction_result = await self._extract_text(file_content, filename, mime_type)
            if not extraction_result['success']:
                return {
                    'success': False,
                    'error': f"Text extraction failed: {extraction_result['error']}",
                    'file_id': file_id
                }
            
            # Prepare document metadata
            doc_metadata = {
                'file_id': file_id,
                'filename': filename,
                'file_hash': file_hash,
                'mime_type': mime_type,
                'file_size': len(file_content),
                'extraction_method': extraction_result.get('method', 'unknown'),
                'upload_timestamp': self._get_timestamp(),
                **(metadata or {})
            }
            
            # Save file to disk (optional)
            file_path = None
            if self.config.get('file_upload', {}).get('save_to_disk', True):
                file_path = await self._save_file_to_disk(file_content, file_id, filename)
                doc_metadata['file_path'] = str(file_path)
            
            # Store in vector database if configured
            storage_results = []
            if self.auto_store and self.retriever_service:
                storage_result = await self._store_in_vector_db(
                    extraction_result['text'],
                    doc_metadata,
                    collection_name
                )
                storage_results.append(storage_result)
            
            return {
                'success': True,
                'file_id': file_id,
                'filename': filename,
                'file_size': len(file_content),
                'mime_type': mime_type,
                'text_length': len(extraction_result['text']),
                'text_preview': extraction_result['text'][:200] + '...' if len(extraction_result['text']) > 200 else extraction_result['text'],
                'metadata': doc_metadata,
                'file_path': str(file_path) if file_path else None,
                'storage_results': storage_results
            }
            
        except Exception as e:
            logger.error(f"Error processing file upload: {str(e)}")
            return {
                'success': False,
                'error': f"File processing failed: {str(e)}",
                'file_id': None
            }
    
    def _validate_file(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """Validate file size and extension."""
        # Check file size
        if len(file_content) > self.max_file_size:
            return {
                'valid': False,
                'error': f"File size ({len(file_content)} bytes) exceeds maximum allowed size ({self.max_file_size} bytes)"
            }
        
        # Check file extension
        file_ext = Path(filename).suffix.lower()
        if self.allowed_extensions and file_ext not in self.allowed_extensions:
            return {
                'valid': False,
                'error': f"File extension '{file_ext}' not allowed. Allowed extensions: {', '.join(self.allowed_extensions)}"
            }
        
        # Check for empty file
        if len(file_content) == 0:
            return {
                'valid': False,
                'error': "File is empty"
            }
        
        return {'valid': True}
    
    def _detect_mime_type(self, file_content: bytes, filename: str) -> str:
        """Detect MIME type of the file."""
        # Try using python-magic if available
        if HAS_MAGIC:
            try:
                mime_type = magic.from_buffer(file_content, mime=True)
                return mime_type
            except:
                pass
        
        # Fallback to mimetypes module
        mime_type, _ = mimetypes.guess_type(filename)
        return mime_type or 'application/octet-stream'
    
    async def _extract_text(self, file_content: bytes, filename: str, mime_type: str) -> Dict[str, Any]:
        """Extract text content from various file types."""
        file_ext = Path(filename).suffix.lower()
        
        try:
            # Text files
            if mime_type.startswith('text/') or file_ext in ['.txt', '.md', '.json']:
                try:
                    text = file_content.decode('utf-8')
                    return {'success': True, 'text': text, 'method': 'utf-8'}
                except UnicodeDecodeError:
                    try:
                        text = file_content.decode('latin-1')
                        return {'success': True, 'text': text, 'method': 'latin-1'}
                    except:
                        return {'success': False, 'error': 'Unable to decode text file'}
            
            # PDF files
            elif mime_type == 'application/pdf' or file_ext == '.pdf':
                return await self._extract_pdf_text(file_content)
            
            # Word documents
            elif mime_type in ['application/vnd.openxmlformats-officedocument.wordprocessingml.document', 
                             'application/msword'] or file_ext in ['.docx', '.doc']:
                return await self._extract_docx_text(file_content)
            
            # Excel files
            elif mime_type in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             'application/vnd.ms-excel'] or file_ext in ['.xlsx', '.xls']:
                return await self._extract_excel_text(file_content)
            
            # CSV files
            elif mime_type == 'text/csv' or file_ext == '.csv':
                return await self._extract_csv_text(file_content)
            
            else:
                return {
                    'success': False,
                    'error': f"Unsupported file type: {mime_type} ({file_ext})"
                }
                
        except Exception as e:
            return {
                'success': False,
                'error': f"Text extraction error: {str(e)}"
            }
    
    async def _extract_pdf_text(self, file_content: bytes) -> Dict[str, Any]:
        """Extract text from PDF files."""
        if not HAS_PDF:
            return {
                'success': False,
                'error': "PDF processing not available. Install pypdf: pip install pypdf"
            }
        
        try:
            pdf_file = io.BytesIO(file_content)
            pdf_reader = PdfReader(pdf_file)
            
            text_parts = []
            for page_num, page in enumerate(pdf_reader.pages):
                try:
                    page_text = page.extract_text()
                    if page_text.strip():
                        text_parts.append(f"Page {page_num + 1}:\n{page_text}")
                except Exception as e:
                    logger.warning(f"Error extracting text from page {page_num + 1}: {str(e)}")
                    continue
            
            if not text_parts:
                return {
                    'success': False,
                    'error': "No text content found in PDF"
                }
            
            return {
                'success': True,
                'text': '\n\n'.join(text_parts),
                'method': 'pypdf'
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f"PDF extraction error: {str(e)}"
            }
    
    async def _extract_docx_text(self, file_content: bytes) -> Dict[str, Any]:
        """Extract text from Word documents."""
        if not HAS_DOCX:
            return {
                'success': False,
                'error': "DOCX processing not available. Install docx2python: pip install docx2python"
            }
        
        try:
            docx_file = io.BytesIO(file_content)
            
            # Use docx2python with context manager for proper resource handling
            with docx2python(docx_file) as doc:
                # Extract all text (docx2python automatically handles paragraphs, tables, etc.)
                extracted_text = doc.text
                
                if not extracted_text or not extracted_text.strip():
                    return {
                        'success': False,
                        'error': "No text content found in Word document"
                    }
                
                return {
                    'success': True,
                    'text': extracted_text,
                    'method': 'docx2python'
                }
            
        except Exception as e:
            return {
                'success': False,
                'error': f"DOCX extraction error: {str(e)}"
            }
    
    async def _extract_excel_text(self, file_content: bytes) -> Dict[str, Any]:
        """Extract text from Excel files."""
        if not HAS_EXCEL:
            return {
                'success': False,
                'error': "Excel processing not available. Install openpyxl: pip install openpyxl"
            }
        
        try:
            excel_file = io.BytesIO(file_content)
            workbook = openpyxl.load_workbook(excel_file, read_only=True)
            
            text_parts = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_parts.append(f"Sheet: {sheet_name}")
                
                for row in sheet.iter_rows(values_only=True):
                    row_values = [str(cell) if cell is not None else '' for cell in row]
                    if any(val.strip() for val in row_values):
                        text_parts.append(' | '.join(row_values))
            
            if len(text_parts) <= len(workbook.sheetnames):  # Only sheet names, no content
                return {
                    'success': False,
                    'error': "No text content found in Excel file"
                }
            
            return {
                'success': True,
                'text': '\n'.join(text_parts),
                'method': 'openpyxl'
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f"Excel extraction error: {str(e)}"
            }
    
    async def _extract_csv_text(self, file_content: bytes) -> Dict[str, Any]:
        """Extract text from CSV files."""
        try:
            # Try different encodings
            for encoding in ['utf-8', 'latin-1', 'cp1252']:
                try:
                    text_content = file_content.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                return {
                    'success': False,
                    'error': "Unable to decode CSV file"
                }
            
            # Parse CSV
            csv_file = io.StringIO(text_content)
            csv_reader = csv.reader(csv_file)
            
            text_parts = []
            for row_num, row in enumerate(csv_reader):
                if row:  # Skip empty rows
                    text_parts.append(' | '.join(str(cell) for cell in row))
            
            if not text_parts:
                return {
                    'success': False,
                    'error': "No content found in CSV file"
                }
            
            return {
                'success': True,
                'text': '\n'.join(text_parts),
                'method': 'csv'
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f"CSV extraction error: {str(e)}"
            }
    
    async def _save_file_to_disk(self, file_content: bytes, file_id: str, filename: str) -> Path:
        """Save uploaded file to disk."""
        file_ext = Path(filename).suffix
        safe_filename = f"{file_id}{file_ext}"
        file_path = self.upload_dir / safe_filename
        
        with open(file_path, 'wb') as f:
            f.write(file_content)
        
        logger.info(f"Saved file to: {file_path}")
        return file_path
    
    async def _store_in_vector_db(self, text: str, metadata: Dict[str, Any], collection_name: str) -> Dict[str, Any]:
        """Store document in vector database."""
        if not self.retriever_service:
            return {
                'success': False,
                'error': "No retriever service available"
            }
        
        try:
            # Ensure the collection exists by setting it on the retriever
            if hasattr(self.retriever_service, 'set_collection'):
                await self.retriever_service.set_collection(collection_name)
                logger.info(f"Set collection to: {collection_name}")
            
            # Split text into chunks if it's too long
            chunks = self._split_text_into_chunks(text)
            
            results = []
            for i, chunk in enumerate(chunks):
                chunk_metadata = metadata.copy()
                chunk_metadata.update({
                    'chunk_index': i,
                    'total_chunks': len(chunks),
                    'chunk_size': len(chunk),
                    'text': chunk,  # Include the actual text content
                    'source_type': 'file_upload'
                })
                
                try:
                    chunk_id = f"{metadata['file_id']}_chunk_{i}"
                    
                    # Store chunk directly in ChromaDB if we have access to the collection
                    if hasattr(self.retriever_service, 'collection') and self.retriever_service.collection:
                        collection = self.retriever_service.collection
                        
                        # Generate embedding for the chunk
                        if hasattr(self.retriever_service, 'embed_query'):
                            embedding = await self.retriever_service.embed_query(chunk)
                            
                            # Clean metadata to ensure no None values
                            clean_metadata = {}
                            for key, value in chunk_metadata.items():
                                if value is None:
                                    clean_metadata[key] = ""
                                else:
                                    clean_metadata[key] = str(value)
                            
                            # Use ChromaDB's upsert method
                            collection.upsert(
                                ids=[chunk_id],
                                embeddings=[embedding],
                                metadatas=[clean_metadata],
                                documents=[chunk]
                            )
                            
                            logger.info(f"Successfully stored chunk {i} in vector database")
                            
                        else:
                            logger.warning("Retriever service doesn't support embedding generation")
                            raise Exception("No embedding method available")
                    else:
                        logger.warning("No ChromaDB collection available in retriever service")
                        raise Exception("No collection available")
                    
                    result = {
                        'success': True,
                        'chunk_id': chunk_id,
                        'chunk_size': len(chunk)
                    }
                    results.append(result)
                    
                except Exception as chunk_error:
                    logger.warning(f"Failed to store chunk {i}: {str(chunk_error)}")
                    result = {
                        'success': False,
                        'chunk_id': f"{metadata['file_id']}_chunk_{i}",
                        'error': str(chunk_error)
                    }
                    results.append(result)
            
            successful_chunks = sum(1 for r in results if r['success'])
            
            return {
                'success': successful_chunks > 0,
                'chunks_stored': successful_chunks,
                'total_chunks': len(results),
                'results': results
            }
            
        except Exception as e:
            logger.error(f"Error storing in vector database: {str(e)}")
            return {
                'success': False,
                'error': f"Vector storage error: {str(e)}"
            }
    
    def _split_text_into_chunks(self, text: str) -> List[str]:
        """Split text into chunks for vector storage."""
        if len(text) <= self.chunk_size:
            return [text]
        
        chunks = []
        start = 0
        
        while start < len(text):
            end = start + self.chunk_size
            
            # If this is not the last chunk, try to find a good break point
            if end < len(text):
                # Look for sentence breaks near the end
                for i in range(end, max(start + self.chunk_size - 200, start), -1):
                    if text[i] in '.!?':
                        end = i + 1
                        break
                else:
                    # Look for word breaks
                    for i in range(end, max(start + self.chunk_size - 100, start), -1):
                        if text[i].isspace():
                            end = i
                            break
            
            chunk = text[start:end].strip()
            if chunk:
                chunks.append(chunk)
            
            # Move start position with overlap
            start = max(end - self.chunk_overlap, start + 1)
        
        return chunks
    
    def _get_timestamp(self) -> str:
        """Get current timestamp in ISO format."""
        from datetime import datetime, timezone
        return datetime.now(timezone.utc).isoformat().replace('+00:00', 'Z')
    
    async def get_file_info(self, file_id: str) -> Optional[Dict[str, Any]]:
        """Get information about an uploaded file."""
        # This would typically query a database
        # For now, we'll check if the file exists on disk
        try:
            file_pattern = f"{file_id}.*"
            matching_files = list(self.upload_dir.glob(file_pattern))
            
            if matching_files:
                file_path = matching_files[0]
                stat = file_path.stat()
                
                return {
                    'file_id': file_id,
                    'file_path': str(file_path),
                    'file_size': stat.st_size,
                    'created_at': stat.st_ctime,
                    'modified_at': stat.st_mtime
                }
            
            return None
            
        except Exception as e:
            logger.error(f"Error getting file info: {str(e)}")
            return None
    
    async def delete_file(self, file_id: str) -> bool:
        """Delete an uploaded file."""
        try:
            file_pattern = f"{file_id}.*"
            matching_files = list(self.upload_dir.glob(file_pattern))
            
            for file_path in matching_files:
                file_path.unlink()
                logger.info(f"Deleted file: {file_path}")
            
            return len(matching_files) > 0
            
        except Exception as e:
            logger.error(f"Error deleting file: {str(e)}")
            return False 